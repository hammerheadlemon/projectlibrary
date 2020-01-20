'''

Store of common functions found in all analysis engine

This is also where the date of BICC is set. This is the date from which much of the analysis is set.
NOTE. Python date format is (YYYY,MM,DD)

'''

import random
import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

'''dates for functions. python date format is Year, Month, day'''
bicc_date = datetime.date(2019, 11, 11)
milestone_analysis_date = datetime.date(2019, 7, 1)

def all_milestone_data_bulk(project_list, master_data):
    '''
    function that filters all milestone data and returns it in dictionary format.

    dictionary is structured as {'project name': {'milestone name': datetime.date: 'notes'}}

    project list: list of project names of interest / in range
    master_data: quarter master data set
    '''

    upper_dictionary = {}

    for name in project_list:
        try:
            p_data = master_data.data[name]
            lower_dictionary = {}
            for i in range(1, 50):
                try:
                    try:
                        lower_dictionary[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast / Actual']: p_data[
                                'Approval MM' + str(i) + ' Notes']}
                    except KeyError:
                        lower_dictionary[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast - Actual']: p_data[
                                'Approval MM' + str(i) + ' Notes']}

                    lower_dictionary[p_data['Assurance MM' + str(i)]] = \
                        {p_data['Assurance MM' + str(i) + ' Forecast - Actual']: p_data[
                                'Assurance MM' + str(i) + ' Notes']}
                except KeyError:
                    pass

            for i in range(18, 67):
                try:
                    lower_dictionary[p_data['Project MM' + str(i)]] = \
                        {p_data['Project MM' + str(i) + ' Forecast - Actual']: p_data['Project MM' + str(i) + ' Notes']}
                except KeyError:
                    pass
        except KeyError:
            lower_dictionary = {}

        upper_dictionary[name] = lower_dictionary

    return upper_dictionary

def approval_milestone_data_bulk(project_list, master_data):
    '''
    function that filters all milestone data and returns it in dictionary format.

    dictionary is structured as {'project name': {'milestone name': datetime.date: 'notes'}}

    project list: list of project names of interest / in range
    master_data: quarter master data set
    '''

    upper_dictionary = {}

    for name in project_list:
        try:
            p_data = master_data.data[name]
            lower_dictionary = {}
            for i in range(1, 50):
                try:
                    try:
                        lower_dictionary[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast / Actual']: p_data[
                                'Approval MM' + str(i) + ' Notes']}
                    except KeyError:
                        lower_dictionary[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast - Actual']: p_data[
                                'Approval MM' + str(i) + ' Notes']}
                except KeyError:
                    pass

        except KeyError:
            lower_dictionary = {}

        upper_dictionary[name] = lower_dictionary

    return upper_dictionary

def ap_p_milestone_data_bulk(project_list, master_data):
    '''
    function that filters  milestone data and returns it in dictionary format.

    dictionary is structured as {'project name': {'milestone name': datetime.date: 'notes'}}

    project list: list of project names of interest / in range
    master_data: quarter master data set
    '''

    upper_dictionary = {}

    for name in project_list:
        try:
            p_data = master_data.data[name]
            lower_dictionary = {}
            for i in range(1, 50):
                try:
                    try:
                        lower_dictionary[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast / Actual'] : p_data['Approval MM' + str(i) + ' Notes']}
                    except KeyError:
                        lower_dictionary[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast - Actual'] : p_data['Approval MM' + str(i) + ' Notes']}

                except KeyError:
                    pass

            for i in range(18, 67):
                try:
                    lower_dictionary[p_data['Project MM' + str(i)]] = \
                        {p_data['Project MM' + str(i) + ' Forecast - Actual'] : p_data['Project MM' + str(i) + ' Notes']}
                except KeyError:
                    pass
        except KeyError:
            lower_dictionary = {}

        upper_dictionary[name] = lower_dictionary

    return upper_dictionary

def assurance_milestone_data_bulk(project_list, master_data):
    """Function to filter out assurance milestone data"""
    upper_dictionary = {}

    for name in project_list:
        try:
            p_data = master_data.data[name]
            lower_dictionary = {}
            for i in range(1, 50):
                lower_dictionary[p_data['Assurance MM' + str(i)]] = \
                    {p_data['Assurance MM' + str(i) + ' Forecast - Actual']: p_data['Assurance MM' + str(i) + ' Notes']}

            upper_dictionary[name] = lower_dictionary
        except KeyError:
            upper_dictionary[name] = {}

    return upper_dictionary

def project_time_difference(proj_m_data_1, proj_m_data_2):
    """Function that calculates time different between milestone dates"""
    upper_dictionary = {}

    for proj_name in proj_m_data_1:
        td_dict = {}
        for milestone in proj_m_data_1[proj_name]:
            if milestone is not None:
                milestone_date = tuple(proj_m_data_1[proj_name][milestone])[0]
                try:
                    if milestone_analysis_date <= milestone_date:
                        try:
                            old_milestone_date = tuple(proj_m_data_2[proj_name][milestone])[0]
                            time_delta = (milestone_date - old_milestone_date).days  # time_delta calculated here
                            if time_delta == 0:
                                td_dict[milestone] = 0
                            else:
                                td_dict[milestone] = time_delta
                        except (KeyError, TypeError):
                            td_dict[milestone] = 'Not reported' # not reported that quarter
                except (KeyError, TypeError):
                    td_dict[milestone] = 'No date provided' # date has now been removed

        upper_dictionary[proj_name] = td_dict

    return upper_dictionary

def filter_gmpp(master):
    project_list = []
    for project_name in master.projects:
        if master.data[project_name]['GMPP - IPA ID Number'] is not None:
            project_list.append(project_name)

    return project_list

def convert_rag_text(dca_rating):

    if dca_rating == 'Green':
        return 'G'
    elif dca_rating == 'Amber/Green':
        return 'A/G'
    elif dca_rating == 'Amber':
        return 'A'
    elif dca_rating == 'Amber/Red':
        return 'A/R'
    elif dca_rating == 'Red':
        return 'R'
    else:
        return ''

def filter_project_group(master_data, group):
    '''
    function for return a list of projects according to their group
    :param master_data: one quarters master data
    :param group: group name of interest. this is a string.
    options are 'Rail Group', 'HSMRPG', 'International Security and Environment', 'Roads Devolution & Motoring'.
    Note this list should be kept up to date as group names change.
    :return: list of projects in specified group
    '''

    project_name_list = master_data.projects

    output_list = []

    for project_name in project_name_list:
        if master_data.data[project_name]['DfT Group'] == group:
            output_list.append(project_name)
        else:
            pass

    return output_list

def get_all_project_names(masters_list):
    '''
    function returns list of all projects across multiple dictionaries

    useful if you need project names across multiple quarters

    masters_list: list of masters containing quarter information
    '''

    output_list = []
    for master in masters_list:
        for name in master.projects:
            if name not in output_list:
                output_list.append(name)

    return output_list

def get_quarter_stamp(masters_list):
    '''
    Function used to specify the quarter being reported.

    masters_list: list of masters containing quarter information
    '''

    output_list = []
    for master in masters_list:
        project_name = random.choice(master.projects)
        quarter_stamp = master.data[project_name]['Reporting period (GMPP - Snapshot Date)']
        output_list.append(quarter_stamp)

    return output_list

def concatenate_dates(date):
    '''
    function for converting dates into concatenated written time periods
    :param date: datetime.date
    :return: concatenated date
    '''
    if date != None:
        a = (date - bicc_date).days
        year = 365
        month = 30
        fortnight = 14
        week = 7
        if a >= 365:
            yrs = int(a / year)
            holding_days_years = a % year
            months = int(holding_days_years / month)
            holding_days_months = a % month
            fortnights = int(holding_days_months / fortnight)
            weeks = int(holding_days_months / week)
        elif 0 <= a <= 365:
            yrs = 0
            months = int(a / month)
            holding_days_months = a % month
            fortnights = int(holding_days_months / fortnight)
            weeks = int(holding_days_months / week)
            # if 0 <= a <=60:
        elif a <= -365:
            yrs = int(a / year)
            holding_days = a % -year
            months = int(holding_days / month)
            holding_days_months = a % -month
            fortnights = int(holding_days_months / fortnight)
            weeks = int(holding_days_months / week)
        elif -365 <= a <= 0:
            yrs = 0
            months = int(a / month)
            holding_days_months = a % -month
            fortnights = int(holding_days_months / fortnight)
            weeks = int(holding_days_months / week)
            # if -60 <= a <= 0:
        else:
            print('something is wrong and needs checking')

        if yrs == 1:
            if months == 1:
                return ('{} yr, {} mth'.format(yrs, months))
            if months > 1:
                return ('{} yr, {} mths'.format(yrs, months))
            else:
                return ('{} yr'.format(yrs))
        elif yrs > 1:
            if months == 1:
                return ('{} yrs, {} mth'.format(yrs, months))
            if months > 1:
                return ('{} yrs, {} mths'.format(yrs, months))
            else:
                return ('{} yrs'.format(yrs))
        elif yrs == 0:
            if a == 0:
                return ('Today')
            elif 1 <= a <= 6:
                return ('This week')
            elif 7 <= a <= 13:
                return ('Next week')
            elif -7 <= a <= -1:
                return ('Last week')
            elif -14 <= a <= -8:
                return ('-2 weeks')
            elif 14 <= a <= 20:
                return ('2 weeks')
            elif 20 <= a <= 60:
                if bicc_date.month == date.month:
                    return ('Later this mth')
                elif (date.month - bicc_date.month) == 1:
                    return ('Next mth')
                else:
                    return ('2 mths')
            elif -60 <= a <= -15:
                if bicc_date.month == date.month:
                    return ('Earlier this mth')
                elif (date.month - bicc_date.month) == -1:
                    return ('Last mth')
                else:
                    return ('-2 mths')
            elif months == 12:
                return ('1 yr')
            else:
                return ('{} mths'.format(months))

        elif yrs == -1:
            if months == -1:
                return ('{} yr, {} mth'.format(yrs, -(months)))
            if months < -1:
                return ('{} yr, {} mths'.format(yrs, -(months)))
            else:
                return ('{} yr'.format(yrs))
        elif yrs < -1:
            if months == -1:
                return ('{} yrs, {} mth'.format(yrs, -(months)))
            if months < -1:
                return ('{} yrs, {} mths'.format(yrs, -(months)))
            else:
                return ('{} yrs'.format(yrs))
    else:
        return ('None')

def up_or_down(latest_dca, last_dca):
    '''
    function that calculates if confidence has increased or decreased
    :param latest_dca:
    :param last_dca:
    :return:
    '''

    if latest_dca == last_dca:
        return (int(0))
    elif latest_dca != last_dca:
        if last_dca == 'Green':
            if latest_dca != 'Amber/Green':
                return (int(-1))
        elif last_dca == 'Amber/Green':
            if latest_dca == 'Green':
                return (int(1))
            else:
                return (int(-1))
        elif last_dca == 'Amber':
            if latest_dca == 'Green':
                return (int(1))
            elif latest_dca == 'Amber/Green':
                return (int(1))
            else:
                return (int(-1))
        elif last_dca == 'Amber/Red':
            if latest_dca == 'Red':
                return (int(-1))
            else:
                return (int(1))
        else:
            return (int(1))

def convert_bc_stage_text(bc_stage):
    '''
    function that converts bc stage.
    :param bc_stage: the string name for business cases that it kept in the master
    :return: standard/shorter string name
    '''

    if bc_stage == 'Strategic Outline Case':
        return 'SOBC'
    elif bc_stage == 'Outline Business Case':
        return 'OBC'
    elif bc_stage == 'Full Business Case':
        return 'FBC'
    elif bc_stage == 'pre-Strategic Outline Case':
        return 'pre-SOBC'
    else:
        return bc_stage

def combine_narrtives(project_name, master_data, key_list):
    '''
    Function that combines narratives across keys
    :param project_name: project name
    :param master_data: master data from one quarter
    :param key_list: list of keys that contain the narrative (values) to be combined.
    :return: combined narrative
    '''
    output = ''
    for key in key_list:
        output = output + str(master_data[project_name][key])

    return output

def baseline_information_bc(project_list, masters_list):
    '''
    Function that calculates when project business case of have changes. Only returns where there have been changes.
    :param project_list: list of project names
    :param masters_list: list of masters with quarter information
    :return: python dictionary in format 'project name':('BC', 'Quarter Stamp', index position in masters_list)
    '''
    output = {}

    for project_name in project_list:
        lower_list = []
        for i, master in enumerate(masters_list):
            if project_name in master.projects:
                approved_bc = master.data[project_name]['BICC approval point']
                quarter = master.data[project_name]['Reporting period (GMPP - Snapshot Date)']
                try:
                    previous_approved_bc = masters_list[i+1].data[project_name]['BICC approval point']
                    if approved_bc != previous_approved_bc:
                        lower_list.append((approved_bc, quarter, i))
                except IndexError:
                    # this captures the last available quarter data if project was in portfolio from beginning
                    lower_list.append((approved_bc, quarter, i))
                except KeyError:
                    # this captures the first quarter the project reported if not in portfolio from beginning
                    lower_list.append((approved_bc, quarter, i))

        output[project_name] = lower_list

    return output

def baseline_information(project_list, masters_list, data_baseline):
    '''
    Function that calculates in information within masters has been baselined
    :param project_list: list of project names
    :param masters_list: list of quarter masters.
    :param data_baseline: type of information to check for baselines. options are: 'ALB milestones' etc
    :return: python dictionary structured as 'project name': ('yes (if reported that quarter), 'quarter stamp',
    index position of master in master quarter list)
    '''

    output = {}

    for project_name in project_list:
        lower_list = []
        for i, master in enumerate(masters_list):
            if project_name in master.projects:
                try:
                    approved_bc = master.data[project_name]['Re-baseline ' + data_baseline]
                    quarter = master.data[project_name]['Reporting period (GMPP - Snapshot Date)']
                    if approved_bc == 'Yes':
                        lower_list.append((approved_bc, quarter, i))
                except KeyError:
                    pass

        output[project_name] = lower_list

    return output

def baseline_index(baseline_data):
    '''
    Function that calculates the index list for baseline data
    :param baseline_data: output created by either baseline_information or baseline_information_bc functions
    :return: python dictionary in format 'project name':[index list]
    '''

    output = {}

    for project_name in baseline_data:
        lower_list = [0, 1]
        for tuple_info in baseline_data[project_name]:
            lower_list.append(tuple_info[2])

        output[project_name] = lower_list

    return output

def get_project_cost_profile(project_name_list, q_masters_data_list, cost_list, year_list, bc_index, index):
    '''
    Function that gets projects project cost information and returns it in a python dictionary format.
    :param project_name_list: list of project names
    :param q_masters_data_list: list of master python dictionaries containing quarter information
    :param cost_list: list of cost key names. this is necessary due to the total cost having be calculated across
    rdel, cdel and non-gov breakdown.
    :param year_list: list of year keys e.g. '19-20', '20-21'
    :param index: index value for which master to use from the q_master_data_list . 0 is for latest, 1 last and
    2 baseline. The actual index list q_master_list is set at a global level in this programme.
    :return: a dictionary structured 'project_name': 'year rdel' : value, 'year cdel' : value, 'year Non-Gov' : value,
    'year total' : value
    '''

    upper_dictionary = {}

    for project_name in project_name_list:
        lower_dictionary = {}
        for year in year_list:
            project_data = q_masters_data_list[bc_index[project_name][index]].data[project_name]
            total = 0
            for type in cost_list:

                try:
                    lower_dictionary[year + type] = project_data[year + type]
                except KeyError:
                    lower_dictionary[year + type] = None

                if year + type in project_data.keys():
                    cost = project_data[year + type]
                    try:
                        total = total + cost
                    except TypeError:
                        pass

            lower_dictionary[year + ' total'] = total

        upper_dictionary[project_name] = lower_dictionary

    return upper_dictionary

def get_project_income_profile(project_name_list, q_masters_data_list, income_list, year_list, bc_index, index):
    '''
    Function that gets projects project income information and returns it in a python dictionary format.
    :param project_name_list: list of project names
    :param q_masters_data_list: list of master python dictionaries containing quarter information
    :param income_list: list of income key names.
    :param year_list: list of year keys e.g. '19-20', '20-21'
    :param index: index value for which master to use from the q_master_data_list . 0 is for latest, 1 last and
    2 baseline. The actual index list q_master_list is set at a global level in this programme.
    :return: a dictionary structured 'project_name' : 'year income' : value
    '''

    upper_dictionary = {}

    for project_name in project_name_list:
        lower_dictionary = {}
        for year in year_list:
            project_data = q_masters_data_list[bc_index[project_name][index]].data[project_name]
            for type in income_list:

                try:
                    lower_dictionary[year + type] = project_data[year + type]
                except KeyError:
                    lower_dictionary[year + type] = 0

        upper_dictionary[project_name] = lower_dictionary

    return upper_dictionary

def calculate_group_project_total(project_name_list, master_data, project_name_no_count_list, type_list, year_list):
    '''
    calculates the total cost figure for each year and type of spend e.g. RDEL 19-20, for all projects of interest.
    :param project_name_list: list of project names
    :param master_data: master data set as created by the get_project_cost_profile
    :param project_name_no_count_list: list of project names to remove from total figures, to ensure no double counting
    e.g. if there are separate schemes as well as overall programme reporting.
    :param type_list: the type of financial figure list being counted. e.g. costs or income
    :return: python dictionary in format 'year + spend type': total
    '''

    output = {}

    project_list = [x for x in project_name_list if x not in project_name_no_count_list]

    for cost in type_list:
        for year in year_list:
            total = 0
            for project_name in project_list:
                try:
                    total = total + master_data[project_name][year + cost]
                except TypeError:
                    total = total + 0

            output[year + cost] = total

    return output

def conditional_formatting(ws):
    '''
    function applies conditional formatting for RAG colors... in development.
    :param ws: worksheet
    :return: worksheet with conditional formatting
    '''

    for column in ws.max_column:
        for i, dca in enumerate(rag_txt_list):
            text = black_text
            fill = fill_colour_list[i]
            dxf = DifferentialStyle(font=text, fill=fill)
            rule = Rule(type="containsText", operator="containsText", text=dca, dxf=dxf)
            for_rule_formula = 'NOT(ISERROR(SEARCH("' + dca + '",' + column + '5)))'
            rule.formula = [for_rule_formula]
            ws.conditional_formatting.add('' + column + '5:' + column + '60', rule)

    return ws

def grey_conditional_formatting(ws):
    '''
    function applies grey conditional formatting for 'Not Reporting'.
    :param worksheet: ws
    :return: cf of sheet
    '''

    grey_text = Font(color="f0f0f0")
    grey_fill = PatternFill(bgColor="f0f0f0")
    dxf = DifferentialStyle(font=grey_text, fill=grey_fill)
    rule = Rule(type="containsText", operator="containsText", text="Not reporting", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Not reporting",A1)))']
    ws.conditional_formatting.add('A1:X80', rule)

    grey_text = Font(color="cfcfea")
    grey_fill = PatternFill(bgColor="cfcfea")
    dxf = DifferentialStyle(font=grey_text, fill=grey_fill)
    rule = Rule(type="containsText", operator="containsText", text="Data not collected", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Data not collected",A1)))']
    ws.conditional_formatting.add('A1:X80', rule)

    return ws

'''old functions not currently in use below here'''

def financial_data(project_list, masters_list, bl_list, cells_to_capture, index):
    '''
    NOT IN USE. OLD FUNCTION. BEING STORED HERE FOR NOW. 

    Function that creates a mini dictionary containing financial information. It used the financial_info function to
    help build the dictionary output.

    project_list: list of project names
    master_list: master data for quarter of interest
    bl_list:
    cells_to_capture: financial info key names. see lists below
    index:

    '''

    output = {}
    for project_name in project_list:
        master_data = masters_list[bl_list[project_name][index]]
        get_financial_info = financial_info(project_name, master_data, cells_to_capture)
        output[project_name] = get_financial_info

    return output

def financial_info(project_name, master_data, cells_to_capture):
    '''
    NOT IN USE. OLD FUNCTION. STORED HERE FOR NOW. 
    function that creates dictionary containing financial {key : value} information.

    project_name = name of project
    master_data = quarter master data set
    cells_to_capture = lists of keys of interest
    '''

    output = {}

    if project_name in master_data.projects:
        for item in master_data.data[project_name]:
            if item in cells_to_capture:
                if master_data.data[project_name][item] is None:
                    output[item] = 0
                else:
                    value = master_data.data[project_name][item]
                    output[item] = value

    else:
        for item in cells_to_capture:
            output[item] = 0

    return output

def calculate_totals(project_name, financial_data):
    '''
    NOT IN USE. OLD FUNCTION. STORED HERE.
    Function that calculates totals.

    project_name: project name
    financial_data: mini project financial dictionary
    '''

    working_data = financial_data[project_name]
    rdel_list = []
    cdel_list = []
    ng_list = []

    for rdel in capture_rdel:
        try:
            rdel_list.append(working_data[rdel])
        except KeyError:
            rdel_list.append(int(0))
    for cdel in capture_cdel:
        try:
            cdel_list.append(working_data[cdel])
        except KeyError:
            cdel_list.append(int(0))
    for ng in capture_ng:
        try:
            ng_list.append(working_data[ng])
        except KeyError:
            ng_list.append(int(0))

    total_list = []
    for i in range(len(rdel_list)):
        total = rdel_list[i] + cdel_list[i] + ng_list[i]
        total_list.append(total)

    return total_list

def calculate_income_totals(project_name, financial_data):
    '''
    FUNCTION NOT IN USE. BEING STORED HERE. 

    function that calculates income totals.

    project_name: project name
    financial_data: mini project financial dictionary
    '''

    working_data = financial_data[project_name]
    income_list = []

    for income in capture_income:
        try:
            income_list.append(working_data[income])
        except KeyError:
            income_list.append(int(0))

    return income_list


def bc_ref_stages(project_list, masters_list):
    """
    NOLONGER IN USE. STORED FOR NOW.
    One of key functions used for calculating which quarter to baseline data from.

    Function returns a dictionary structured in the following way project name[('latest quarter info', 'latest bc'),
    ('last quarter info', 'last bc'), ('last baseline quarter info', 'last baseline bc'), ('oldest quarter info',
    'oldest bc')] depending on the amount information available in the data. Only the first three key values are returned,
    to ensure consistency (which is helpful later).

    project_list: list of project names
    masters_list = list of master dictionaries

    """
    output = {}

    for project_name in project_list:
        # print(project_name)
        all_list = []  # format [('quarter info': 'bc')] across all masters including project
        bl_list = []  # format ['bc', 'bc'] across all masters. bl_list_2 removes duplicates
        ref_list = []  # format as for all list but only contains the three tuples of interest
        for master in masters_list:
            try:
                bc_stage = master.data[project_name]['BICC approval point']
                quarter = master.data[project_name]['Reporting period (GMPP - Snapshot Date)']
                tuple = (quarter, bc_stage)
                all_list.append(tuple)
            except KeyError:
                pass

        for i in range(0, len(all_list)):
            bl_list.append(all_list[i][1])

        '''below lines of text from stackoverflow. Question, remove duplicates in python list while
        preserving order'''
        seen = set()
        seen_add = seen.add
        bl_list_2 = [x for x in bl_list if not (x in seen or seen_add(x))]

        ref_list.insert(0, all_list[0])  # puts the latest info into the list first

        try:
            ref_list.insert(1, all_list[1])  # puts that last info into the list
        except IndexError:
            ref_list.insert(1, all_list[0])

        if len(bl_list_2) == 1:  # puts oldest info into list (as basline if no baseline)
            ref_list.insert(2, all_list[-1])
        else:
            for i in range(0, len(all_list)):  # puts in baseline
                if all_list[i][1] == bl_list[0]:
                    ref_list.insert(2, all_list[i])

        '''there is a hack here i.e. returning only first three in ref_list. There's a bug which I don't fully
        understand, but this solution is hopefully good enough for now'''
        output[project_name] = ref_list[0:3]

    return output


def master_baseline_index(project_list, masters_list, baselines_list):
    """
    NOLONGER IN USE. STORED FOR NOW.

    Another key function to calculate which quarter to baseline data from.

    Function returns a dictionary structured as {'project name': [n,n,n]}.
    The n (number) values denote where the relevant baseline master dictionaries are list of master dictionaries.
    The first n in the latest master, second n is last master, third n is baseline master.

    project_list: list of projects
    masters_list: list of masters
    baseline_list: list of project baseline information in the structure {'project name': [('quarter stamp', 'bc stage),
    (), ()] as created by bc_ref_stage function.

    """
    output = {}

    for project_name in project_list:
        master_q_list = []
        for key in baselines_list[project_name]:
            for x, master in enumerate(masters_list):
                try:
                    quarter = master.data[project_name]['Reporting period (GMPP - Snapshot Date)']
                    if quarter == key[0]:
                        master_q_list.append(x)
                except KeyError:
                    pass

        output[project_name] = master_q_list

    return output